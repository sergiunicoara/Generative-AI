"""Local relational-to-KG ingestion with declarative, tenant-scoped mappings.

This module deliberately has no cloud or vendor dependency.  SQLite is used as
the local reference connector; the mapping and validation contracts are
provider-neutral and can later be backed by another database adapter.
"""

from __future__ import annotations

import asyncio
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, uuid5

from pydantic import BaseModel, Field, model_validator

from graphrag.core.models import Chunk, Document, Entity, Relation, SourceType
from graphrag.graph.source_catalog import (
    SourceCatalogRepository,
    SourceEnvelope,
    SourceKind,
    SourceMapping,
    SourceSystem,
)

_IDENTIFIER = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _identifier(value: str, label: str) -> str:
    if not _IDENTIFIER.fullmatch(value):
        raise ValueError(f"{label} must be a simple SQL identifier")
    return value


class EntityTableMapping(BaseModel):
    table: str = Field(min_length=1)
    entity_type: str = Field(min_length=1)
    id_column: str = Field(min_length=1)
    name_column: str = Field(min_length=1)
    description_column: str | None = None
    attributes: list[str] = Field(default_factory=list)

    @model_validator(mode="after")
    def validate_identifiers(self) -> "EntityTableMapping":
        for key in ("table", "id_column", "name_column", "description_column"):
            value = getattr(self, key)
            if value:
                _identifier(value, key)
        for value in self.attributes:
            _identifier(value, "attributes")
        return self


class RelationTableMapping(BaseModel):
    table: str = Field(min_length=1)
    source_table: str = Field(min_length=1)
    target_table: str = Field(min_length=1)
    source_column: str = Field(min_length=1)
    target_column: str = Field(min_length=1)
    relation: str = Field(min_length=2, pattern=r"^[A-Z][A-Z0-9_]{1,49}$")
    confidence_column: str | None = None
    valid_from_column: str | None = None
    valid_to_column: str | None = None

    @model_validator(mode="after")
    def validate_identifiers(self) -> "RelationTableMapping":
        for key in (
            "table", "source_table", "target_table", "source_column", "target_column", "confidence_column",
            "valid_from_column", "valid_to_column",
        ):
            value = getattr(self, key)
            if value:
                _identifier(value, key)
        return self


class RelationalGraphMapping(BaseModel):
    """Declarative mapping from tables to KG entities and relations."""

    id: str = Field(min_length=1)
    version: str = Field(min_length=1)
    source_id: str = Field(min_length=1)
    tenant: str = Field(min_length=1)
    entities: list[EntityTableMapping] = Field(min_length=1)
    relations: list[RelationTableMapping] = Field(default_factory=list)
    ontology_version: str = "local/v1"

    def as_source_mapping(self) -> SourceMapping:
        return SourceMapping(
            id=f"{self.id}-{self.version}",
            tenant=self.tenant,
            source_id=self.source_id,
            version=self.version,
            mapping=self.model_dump(mode="json", exclude={"tenant", "source_id"}),
        )


class MappingValidationReport(BaseModel):
    valid: bool
    tenant: str
    source_id: str
    entity_rows: int = 0
    relation_rows: int = 0
    errors: list[str] = Field(default_factory=list)
    shacl_conforms: bool | None = None


class SQLiteSourceConnector:
    """Read-only local SQLite connector implementing the source contract."""

    kind = SourceKind.DATABASE

    def __init__(self, path: str | Path):
        self.path = Path(path)

    @property
    def uri(self) -> str:
        return str(self.path.resolve())

    async def records(
        self, source: SourceSystem, mapping: SourceMapping, *, cursor: str = ""
    ):
        spec = mapping.mapping
        tables = list(spec.get("entities", [])) + list(spec.get("relations", []))
        for table_spec in tables:
            table = _identifier(str(table_spec["table"]), "table")
            rows = await self.read_table(table)
            for index, row in enumerate(rows):
                external_id = f"{table}:{row.get('id', index)}"
                yield SourceEnvelope(
                    external_id=external_id,
                    content=json.dumps(row, sort_keys=True, default=str),
                    content_type="application/json",
                    metadata={"table": table, "source_id": source.id},
                    cursor=str(index + 1),
                )

    def _read_table(self, table: str) -> list[dict[str, Any]]:
        if not self.path.exists():
            raise FileNotFoundError(self.path)
        with sqlite3.connect(f"file:{self.path}?mode=ro", uri=True) as conn:
            conn.row_factory = sqlite3.Row
            return [dict(row) for row in conn.execute(f'SELECT * FROM "{table}"')]

    async def read_table(self, table: str) -> list[dict[str, Any]]:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._read_table, _identifier(table, "table"))


class PostgreSQLSourceConnector:
    """Read-only PostgreSQL adapter for the same local mapping contract.

    It works with a local Docker PostgreSQL instance through a SQLAlchemy
    async URL. Table names are identifier-validated; all values remain data,
    never executable SQL supplied by a mapping or an agent.
    """

    kind = SourceKind.DATABASE

    def __init__(self, url: str):
        if not url.startswith(("postgresql+asyncpg://", "postgresql://")):
            raise ValueError("PostgreSQL URL must use postgresql+asyncpg:// or postgresql://")
        self.url = url
        self._engine = None

    @property
    def uri(self) -> str:
        """Safe source identity; credentials never enter a persisted KGSource."""
        from sqlalchemy.engine import make_url

        return str(make_url(self.url).render_as_string(hide_password=True))

    def _get_engine(self):
        if self._engine is None:
            from sqlalchemy.ext.asyncio import create_async_engine
            self._engine = create_async_engine(self.url, pool_pre_ping=True)
        return self._engine

    async def read_table(self, table: str) -> list[dict[str, Any]]:
        from sqlalchemy import text

        safe_table = _identifier(table, "table")
        async with self._get_engine().connect() as connection:
            result = await connection.execute(text(f'SELECT * FROM "{safe_table}"'))
            return [dict(row) for row in result.mappings().all()]

    async def records(
        self, source: SourceSystem, mapping: SourceMapping, *, cursor: str = ""
    ):
        spec = mapping.mapping
        tables = list(spec.get("entities", [])) + list(spec.get("relations", []))
        for table_spec in tables:
            table = _identifier(str(table_spec["table"]), "table")
            for index, row in enumerate(await self.read_table(table)):
                yield SourceEnvelope(
                    external_id=f"{table}:{row.get('id', index)}",
                    content=json.dumps(row, sort_keys=True, default=str),
                    content_type="application/json",
                    metadata={"table": table, "source_id": source.id},
                    cursor=str(index + 1),
                )

    async def close(self) -> None:
        if self._engine is not None:
            await self._engine.dispose()
            self._engine = None


class ExcelWorkbookConnector:
    """Read an Excel workbook through the same mapping contract as databases.

    A worksheet is treated as a table and its first row as identifier-safe
    column names.  This keeps spreadsheet ingestion declarative, SHACL-gated,
    and provenance-preserving instead of introducing a separate ad-hoc path.
    """

    kind = SourceKind.FILE

    def __init__(self, path: str | Path):
        self.path = Path(path)

    @property
    def uri(self) -> str:
        return self.path.resolve().as_uri()

    def _read_table(self, sheet_name: str) -> list[dict[str, Any]]:
        if not self.path.exists():
            raise FileNotFoundError(self.path)
        from openpyxl import load_workbook

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"worksheet not found: {sheet_name}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        names = [str(header or "").strip() for header in headers]
        if not all(names) or len(set(names)) != len(names):
            raise ValueError(f"{sheet_name}: headers must be non-empty and unique")
        for name in names:
            _identifier(name, f"{sheet_name} header")
        return [
            dict(zip(names, row, strict=True))
            for row in rows
            if any(value is not None and value != "" for value in row)
        ]

    async def read_table(self, table: str) -> list[dict[str, Any]]:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._read_table, table)

    async def records(
        self, source: SourceSystem, mapping: SourceMapping, *, cursor: str = ""
    ):
        spec = mapping.mapping
        sheets = list(spec.get("entities", [])) + list(spec.get("relations", []))
        for sheet_spec in sheets:
            sheet = str(sheet_spec["table"])
            for index, row in enumerate(await self.read_table(sheet)):
                yield SourceEnvelope(
                    external_id=f"{sheet}:{row.get('id', index)}",
                    content=json.dumps(row, sort_keys=True, default=str),
                    content_type="application/json",
                    metadata={"worksheet": sheet, "source_id": source.id},
                    cursor=str(index + 1),
                )


class RelationalGraphIngestor:
    """Validate and persist mapped relational rows through ``GraphWriter``."""

    def __init__(self, connector: SQLiteSourceConnector | PostgreSQLSourceConnector | ExcelWorkbookConnector, graph_writer):
        self.connector = connector
        self.graph_writer = graph_writer

    async def validate(self, mapping: RelationalGraphMapping) -> MappingValidationReport:
        errors: list[str] = []
        entity_keys: set[tuple[str, str]] = set()
        entity_rows = 0
        relation_rows = 0

        for table_map in mapping.entities:
            rows = await self.connector.read_table(table_map.table)
            entity_rows += len(rows)
            for row in rows:
                key = row.get(table_map.id_column)
                name = row.get(table_map.name_column)
                if key in (None, "") or name in (None, ""):
                    errors.append(
                        f"{table_map.table}: every row needs {table_map.id_column} and {table_map.name_column}"
                    )
                else:
                    entity_keys.add((table_map.table, str(key)))

        for table_map in mapping.relations:
            rows = await self.connector.read_table(table_map.table)
            relation_rows += len(rows)
            for row in rows:
                if row.get(table_map.source_column) in (None, ""):
                    errors.append(f"{table_map.table}: missing {table_map.source_column}")
                if row.get(table_map.target_column) in (None, ""):
                    errors.append(f"{table_map.table}: missing {table_map.target_column}")

        return MappingValidationReport(
            valid=not errors,
            tenant=mapping.tenant,
            source_id=mapping.source_id,
            entity_rows=entity_rows,
            relation_rows=relation_rows,
            errors=sorted(set(errors)),
        )

    async def ingest(self, mapping: RelationalGraphMapping) -> MappingValidationReport:
        report = await self.validate(mapping)
        if not report.valid:
            raise ValueError("relational mapping rejected: " + "; ".join(report.errors))

        entities: list[Entity] = []
        by_source_key: dict[tuple[str, str], Entity] = {}
        payload: list[dict[str, Any]] = []
        for table_map in mapping.entities:
            rows = await self.connector.read_table(table_map.table)
            for row in rows:
                source_key = str(row[table_map.id_column])
                entity_id = str(uuid5(NAMESPACE_URL, f"{mapping.source_id}:{table_map.table}:{source_key}"))
                entity = Entity(
                    id=entity_id,
                    name=str(row[table_map.name_column]),
                    type=table_map.entity_type,
                    description=str(row.get(table_map.description_column) or "")
                    if table_map.description_column else "",
                    tenant=mapping.tenant,
                    source_type=SourceType.DOCUMENT,
                    source_doc_id=f"relational:{mapping.source_id}",
                )
                entities.append(entity)
                by_source_key[(table_map.table, source_key)] = entity
                payload.append({"table": table_map.table, "row": row})

        relations: list[Relation] = []
        for table_map in mapping.relations:
            for row in await self.connector.read_table(table_map.table):
                source = by_source_key.get((table_map.source_table, str(row[table_map.source_column])))
                target = by_source_key.get((table_map.target_table, str(row[table_map.target_column])))
                if source is None or target is None:
                    raise ValueError(f"{table_map.table}: relation references an unknown entity")
                relation_id = str(uuid5(NAMESPACE_URL, f"{mapping.source_id}:{table_map.table}:{source.id}:{target.id}"))
                relations.append(Relation(
                    id=relation_id,
                    source_entity_id=source.id,
                    target_entity_id=target.id,
                    relation=table_map.relation,
                    confidence=float(row.get(table_map.confidence_column, 1.0) or 1.0)
                    if table_map.confidence_column else 1.0,
                    valid_from=self._timestamp(row.get(table_map.valid_from_column)) if table_map.valid_from_column else None,
                    valid_to=self._timestamp(row.get(table_map.valid_to_column)) if table_map.valid_to_column else None,
                    source_doc_id=f"relational:{mapping.source_id}",
                ))

        from graphrag.graph.shacl_validator import SHACLValidator

        conforms, shacl_report = SHACLValidator.validate_relational_batch(
            entities, relations, tenant=mapping.tenant,
        )
        report.shacl_conforms = conforms
        if not conforms:
            raise ValueError("relational mapping rejected by SHACL: " + shacl_report)

        # A Document.source_id is a real foreign-key-like graph contract: make
        # the source and immutable mapping version durable before the document
        # write so `INGESTED_FROM` can be formed atomically by merge_document.
        catalog = SourceCatalogRepository(self.graph_writer.neo4j_client)
        await catalog.upsert_source(SourceSystem(
            id=mapping.source_id,
            tenant=mapping.tenant,
            name=mapping.source_id,
            kind=self.connector.kind,
            uri=self.connector.uri,
            owner="relational-ingestion",
            classification="synthetic" if mapping.tenant == "sustainability" else "internal",
        ))
        await catalog.add_mapping(mapping.as_source_mapping())

        raw = json.dumps(payload, sort_keys=True, default=str)
        document = Document(
            id=str(uuid5(NAMESPACE_URL, f"relational-document:{mapping.source_id}:{mapping.version}")),
            filename=f"relational://{mapping.source_id}/{mapping.version}",
            source_path=self.connector.uri,
            raw_text=raw,
            tenant=mapping.tenant,
            source_id=mapping.source_id,
            status="done",
            metadata={"mapping_id": mapping.id, "mapping_version": mapping.version,
                      "ontology_version": mapping.ontology_version,
                      "provenance": "local-relational-source"},
        )
        document_id = await self.graph_writer.write_document(document)
        chunk = Chunk(
            id=str(uuid5(NAMESPACE_URL, f"relational-chunk:{document_id}")),
            document_id=document_id,
            text=raw,
            chunk_index=0,
            tenant=mapping.tenant,
            metadata={"source_id": mapping.source_id, "mapping_version": mapping.version},
        )
        await self.graph_writer.write_chunks([chunk])
        written = await self.graph_writer.write_entities(entities, chunk)
        entity_map = {entity.id: entity for entity in written}
        await self.graph_writer.write_relations(relations, entity_map, doc_id=document_id, tenant=mapping.tenant)
        return report

    @staticmethod
    def _timestamp(value: Any) -> datetime | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).astimezone(timezone.utc)


__all__ = [
    "EntityTableMapping", "RelationTableMapping", "RelationalGraphMapping",
    "MappingValidationReport", "SQLiteSourceConnector", "PostgreSQLSourceConnector",
    "ExcelWorkbookConnector", "RelationalGraphIngestor",
]
